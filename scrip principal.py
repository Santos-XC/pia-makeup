import requests
import json
import matplotlib.pyplot as plt
from collections import Counter 

# Función para hacer una solicitud a la API
def consultar_api(api_url):
    response = requests.get(api_url)
    if response.status_code == 200:
        return response.json()
    else:
        print("Error al obtener datos de la API")
        return None

# Función para guardar los datos en un archivo
def guardar_datos(data, archivo):
    with open(archivo, "w") as file:
        json.dump(data, file)
    print(f"Datos guardados en {archivo}")

# Función para cargar datos desde un archivo
def cargar_datos(archivo):
    try:
        with open(archivo, "r") as file:
            data = json.load(file)
            return data
    except FileNotFoundError:
        print(f"El archivo {archivo} no existe.")
        return None

# Función para realizar análisis estadístico
def analizar_datos(data):
    if not data:
        return None

    # Calcular el promedio de los precios
    precios = [float(producto["price"]) for producto in data if producto.get("price") is not None]
    promedio = sum(precios) / len(precios)

    # Calcular el producto más comprado
    productos = [producto["name"] for producto in data if "name" in producto]
    producto_mas_comprado = Counter(productos).most_common(1)[0][0]

    # Calcular el maquillaje más utilizado
    tipos_maquillaje = [producto["category"] for producto in data if "category" in producto and producto["category"] is not None]
    maquillaje_mas_utilizado = Counter(tipos_maquillaje).most_common(1)[0][0]

    # Crear archivo Excel con los datos
    crear_excel(data, "makeup_data.xlsx")

    return promedio, producto_mas_comprado, maquillaje_mas_utilizado 


import openpyxl
# Función para crear un archivo Excel
def crear_excel(data, excel_file):
    if not data:
        return

    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir los encabezados
    headers = list(data[0].keys()) if data else []
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Escribir los datos
    for row_num, producto in enumerate(data, start=2):
        for col_num, value in enumerate(producto.values(), start=1):
            # Convertir listas a cadenas antes de escribirlas en Excel
            if isinstance(value, list):
                value = ', '.join(map(str, value))
            sheet.cell(row=row_num, column=col_num, value=value)

    # Guardar el libro de Excel
    workbook.save(excel_file)
    print(f"Datos guardados en {excel_file}")

# Función para generar gráficas
def generar_graficas(data):
    if not data:
        return

    # Generar histograma de precios
    precios = [float(producto["price"]) for producto in data if producto.get("price") is not None]
    plt.hist(precios, bins=20, edgecolor="k")
    plt.xlabel("Precio")
    plt.ylabel("Frecuencia")
    plt.title("Histograma de Precios")
    plt.show()

    # Generar gráfica de productos más comprados
    productos = [producto["name"] for producto in data if "name" in producto]
    contar_productos = Counter(productos)
    productos_comunes = contar_productos.most_common(5) 
    nombres_productos, frecuencias = zip(*productos_comunes)
    plt.bar(nombres_productos, frecuencias, color='skyblue')
    plt.xlabel("Producto")
    plt.ylabel("Frecuencia")
    plt.title("Productos Más Comprados")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

    # Generar gráfica de tipos de maquillaje más utilizados
    tipos_maquillaje = [producto["category"] for producto in data if "category" in producto and producto["category"] is not None]
    contar_tipos_maquillaje = Counter(tipos_maquillaje)
    tipos_maquillaje_comunes = contar_tipos_maquillaje.most_common()

    nombres_tipos, frecuencias_tipos = zip(*tipos_maquillaje_comunes)
    plt.bar(nombres_tipos, frecuencias_tipos, color='salmon')
    plt.xlabel("Tipo de Maquillaje")
    plt.ylabel("Frecuencia")
    plt.title("Tipos de Maquillaje Más Utilizados")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

# URL de la API makeup-api
api_url = "http://makeup-api.herokuapp.com/api/v1/products.json"

#inicializacion de datos
data = None

while True:
    opcion = input("¿Qué deseas hacer?\n1. Consultar API\n2. Cargar datos\n3. Analizar datos\n4. Generar gráficas\n5. Salir\nOpción: ")

    if opcion == "1":
        data = consultar_api(api_url)
        if data:
            guardar_opcion = input("¿Deseas guardar los datos consultados? (S/N): ").strip().lower()
            if guardar_opcion == "s":
                guardar_datos(data, "makeup_data.json")
    elif opcion == "2":
        data = cargar_datos("makeup_data.json")
    elif opcion == "3":
        resultado_analisis = analizar_datos(data)

        if resultado_analisis:
            promedio, producto_mas_comprado, maquillaje_mas_utilizado = resultado_analisis
            print(f"Promedio de precios: ${promedio:.2f}")
        else:
            print("No hay datos para analizar. Consulta la API o carga datos.")
            
    elif opcion == "4":
        generar_graficas(data)
    elif opcion == "5":
        break
    else:
        print("Opción no válida. Introduce una opción válida.")